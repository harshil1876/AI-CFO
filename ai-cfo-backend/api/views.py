from rest_framework import generics, status
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from .models import (
    UploadedFile, ParsedRecord,
    Transaction, DepartmentData,
    KPISnapshot, ForecastResult, AnomalyLog, Recommendation,
    DataSource, ConnectorSyncLog, Budget, PurchaseOrder, Invoice, NotificationMeta,
    Workspace, GoalTarget, OrgChatMessage,
    InvoiceThread, InboundEmailLog
)
from django.utils import timezone
import os
from rest_framework.exceptions import PermissionDenied

def get_verified_bot_id(request, provided_bot_id):
    """
    Validates that the requested bot_id matches the Clerk JWT's org_id or user_id.
    """
    if not provided_bot_id:
        return provided_bot_id
        
    clerk_org = getattr(request, 'clerk_org_id', None)
    clerk_user = getattr(request, 'clerk_user_id', None)
    
    # Development fallback (bypass strict check locally)
    is_debug = str(os.environ.get("DEBUG", "False")).lower() == "true"
    host = request.get_host() if hasattr(request, 'get_host') else ""
    if is_debug or 'localhost' in host or '127.0.0.1' in host:
        return provided_bot_id

    # The provided bot_id must match either the Clerk active Organization ID or User ID
    if provided_bot_id != clerk_org and provided_bot_id != clerk_user:
        raise PermissionDenied("Unauthorized: Workspace access denied.")
        
    return provided_bot_id

from .serializers import (
    UploadedFileSerializer, ParsedRecordSerializer,
    TransactionSerializer, DepartmentDataSerializer,
    KPISnapshotSerializer, ForecastResultSerializer,
    AnomalyLogSerializer, RecommendationSerializer, BudgetSerializer,
    PurchaseOrderSerializer, InvoiceSerializer,
    WorkspaceSerializer, GoalTargetSerializer, OrgChatMessageSerializer
)
from .services.descriptive_analytics import calculate_kpis
from .services.forecasting import run_revenue_forecast
from .services.anomaly_detection import detect_anomalies
from .services.prescriptive_logic import generate_recommendations
from .services.file_processor import save_uploaded_file, process_file
from .services.rag_engine import sync_financial_context_to_rag
from .services.chat_service import chat_with_cfo
from .services.simulation_engine import run_simulation
from .services.alert_service import check_and_send_alerts
from .services.budgeting_engine import calculate_variance, run_monte_carlo_simulation


# ──────────────────────────────────────────────
# Sprint 1: Flexible File Upload Endpoints
# ──────────────────────────────────────────────

@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
def upload_file(request):
    """
    Upload any financial file (CSV, Excel, JSON, etc.)
    The system will auto-detect the schema and generate an AI summary.
    """
    bot_id = get_verified_bot_id(request, request.data.get("bot_id"))
    file_obj = request.FILES.get("file")

    if not bot_id:
        return Response(
            {"error": "bot_id is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not file_obj:
        return Response(
            {"error": "No file provided. Use 'file' field in multipart form."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Save the file
    upload = save_uploaded_file(file_obj, bot_id)

    # Process the file (parse + AI analysis)
    result = process_file(upload.id)

    return Response(result, status=status.HTTP_201_CREATED if result["status"] == "completed" else status.HTTP_422_UNPROCESSABLE_ENTITY)


class UploadedFileListView(generics.ListAPIView):
    """List all uploaded files for a bot."""
    serializer_class = UploadedFileSerializer

    def get_queryset(self):
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get("bot_id"))
        if bot_id:
            return UploadedFile.objects.filter(bot_id=bot_id).order_by("-uploaded_at")
        return UploadedFile.objects.all().order_by("-uploaded_at")


class ParsedRecordListView(generics.ListAPIView):
    """List parsed records for a specific uploaded file."""
    serializer_class = ParsedRecordSerializer

    def get_queryset(self):
        file_id = self.request.query_params.get("file_id")
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get("bot_id"))
        qs = ParsedRecord.objects.all()
        if file_id:
            qs = qs.filter(source_file_id=file_id)
        if bot_id:
            qs = qs.filter(bot_id=bot_id)
        return qs.order_by("row_index")


# ──────────────────────────────────────────────
# Sprint 1: Legacy Structured Endpoints (kept for backward compat)
# ──────────────────────────────────────────────

class TransactionListCreateView(generics.ListCreateAPIView):
    serializer_class = TransactionSerializer

    def get_queryset(self):
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get('bot_id'))
        if bot_id:
            return Transaction.objects.filter(bot_id=bot_id)
        return Transaction.objects.all()


class DepartmentDataListCreateView(generics.ListCreateAPIView):
    serializer_class = DepartmentDataSerializer

    def get_queryset(self):
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get('bot_id'))
        if bot_id:
            return DepartmentData.objects.filter(bot_id=bot_id)
        return DepartmentData.objects.all()


# ──────────────────────────────────────────────
# Sprint 17: Transaction Review Status PATCH
# ──────────────────────────────────────────────

@api_view(['PATCH'])
def update_transaction_status(request, transaction_id):
    """
    PATCH /api/transactions/<id>/status/
    Body: { "review_status": "reviewed" | "flagged" | "approved" | "pending" }
    """
    bot_id = get_verified_bot_id(request, request.query_params.get('bot_id') or request.data.get('bot_id'))
    new_status = request.data.get('review_status')
    VALID_STATUSES = ['pending', 'reviewed', 'flagged', 'approved']

    if not new_status or new_status not in VALID_STATUSES:
        return Response(
            {'error': f'review_status must be one of: {VALID_STATUSES}'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        qs = Transaction.objects.filter(pk=transaction_id)
        if bot_id:
            qs = qs.filter(bot_id=bot_id)
        txn = qs.get()
        txn.review_status = new_status
        txn.save(update_fields=['review_status'])
        return Response({'id': txn.id, 'review_status': txn.review_status})
    except Transaction.DoesNotExist:
        return Response({'error': 'Transaction not found.'}, status=status.HTTP_404_NOT_FOUND)


# ──────────────────────────────────────────────
# Sprint 17: Usage Metrics Endpoint
# ──────────────────────────────────────────────

@api_view(['GET'])
def usage_metrics(request):
    """
    GET /api/usage/?bot_id=xxx
    Returns real org-level usage stats: file uploads, rows ingested, transactions, chat queries.
    """
    import os
    bot_id = get_verified_bot_id(request, request.query_params.get('bot_id'))
    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    file_count = UploadedFile.objects.filter(bot_id=bot_id).count()
    total_rows = UploadedFile.objects.filter(bot_id=bot_id).values_list('row_count', flat=True)
    rows_ingested = sum(total_rows)
    transaction_count = Transaction.objects.filter(bot_id=bot_id).count()

    # Approximate DB space: avg 500 bytes per row
    db_bytes = rows_ingested * 500
    db_mb = round(db_bytes / (1024 * 1024), 2)

    return Response({
        'file_uploads': file_count,
        'rows_ingested': rows_ingested,
        'transaction_count': transaction_count,
        'db_storage_mb': db_mb,
        'db_storage_display': f"{db_mb} MB" if db_mb < 1000 else f"{round(db_mb/1024, 2)} GB",
    })


# ──────────────────────────────────────────────
# Sprint 2: Intelligence Engine Endpoints
# ──────────────────────────────────────────────

@api_view(["POST"])
def run_analytics(request):
    """
    Trigger full analytics pipeline for a bot:
    1. Calculate KPIs
    2. Detect anomalies
    3. Generate prescriptive recommendations
    """
    bot_id = get_verified_bot_id(request, request.data.get("bot_id"))
    period = request.data.get("period")

    if not bot_id or not period:
        return Response(
            {"error": "bot_id and period are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    kpis = calculate_kpis(bot_id, period)
    anomalies = detect_anomalies(bot_id)
    recommendations = generate_recommendations(bot_id, period)

    return Response({
        "status": "success",
        "kpis": kpis,
        "anomalies_found": len(anomalies),
        "anomalies": anomalies,
        "recommendations": recommendations,
    })


@api_view(["POST"])
def run_forecast(request):
    """Trigger Prophet revenue forecast for a bot."""
    bot_id = get_verified_bot_id(request, request.data.get("bot_id"))
    periods = request.data.get("periods", 6)

    if not bot_id:
        return Response(
            {"error": "bot_id is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    forecasts = run_revenue_forecast(bot_id, periods=int(periods))

    return Response({
        "status": "success",
        "bot_id": bot_id,
        "forecast_months": len(forecasts),
        "forecasts": forecasts,
    })


# ──────────────────────────────────────────────
# Sprint 2: Read-only List Endpoints
# ──────────────────────────────────────────────

class KPISnapshotListView(generics.ListAPIView):
    serializer_class = KPISnapshotSerializer

    def get_queryset(self):
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get("bot_id"))
        if bot_id:
            return KPISnapshot.objects.filter(bot_id=bot_id).order_by("-created_at")
        return KPISnapshot.objects.all().order_by("-created_at")


class ForecastResultListView(generics.ListAPIView):
    serializer_class = ForecastResultSerializer

    def get_queryset(self):
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get("bot_id"))
        if bot_id:
            return ForecastResult.objects.filter(bot_id=bot_id).order_by("forecast_date")
        return ForecastResult.objects.all().order_by("forecast_date")


class AnomalyLogListView(generics.ListAPIView):
    serializer_class = AnomalyLogSerializer

    def get_queryset(self):
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get("bot_id"))
        if bot_id:
            return AnomalyLog.objects.filter(bot_id=bot_id).order_by("-detected_at")
        return AnomalyLog.objects.all().order_by("-detected_at")


class RecommendationListView(generics.ListAPIView):
    serializer_class = RecommendationSerializer

    def get_queryset(self):
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get("bot_id"))
        if bot_id:
            return Recommendation.objects.filter(bot_id=bot_id).order_by("-created_at")
        return Recommendation.objects.all().order_by("-created_at")


# ──────────────────────────────────────────────
# Sprint 3: Conversational CFO & RAG Endpoints
# ──────────────────────────────────────────────

@api_view(["POST"])
def chat(request, bot_id):
    """
    AI CFO Chat endpoint — the core conversational interface.
    Retrieves RAG context + live KPIs, then queries Gemini LLM.
    """
    message = request.data.get("message")
    chat_history = request.data.get("history", [])
    agent_id = request.data.get("agent_id", "strategist")

    if not message:
        return Response(
            {"error": "message is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    result = chat_with_cfo(bot_id, message, chat_history, agent_id)

    return Response(result)


@api_view(["POST"])
def sync_rag(request):
    """
    Trigger RAG sync: push latest financial intelligence to Upstash Search.
    Should be called after analytics pipeline completes.
    """
    bot_id = request.data.get("bot_id")

    if not bot_id:
        return Response(
            {"error": "bot_id is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    result = sync_financial_context_to_rag(bot_id)

    return Response({
        "status": "success",
        **result,
    })


# ──────────────────────────────────────────────
# Sprint 4: Simulation & Alert Endpoints
# ──────────────────────────────────────────────

@api_view(["POST"])
def simulate(request):
    """
    Run a what-if simulation on financial data.
    Accepts hypothetical changes and returns projected KPIs.
    """
    bot_id = request.data.get("bot_id")
    period = request.data.get("period")
    scenarios = request.data.get("scenarios", [])

    if not bot_id or not period:
        return Response(
            {"error": "bot_id and period are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not scenarios:
        return Response(
            {"error": "At least one scenario is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    result = run_simulation(bot_id, period, scenarios)

    if "error" in result:
        return Response(result, status=status.HTTP_404_NOT_FOUND)

    return Response({"status": "success", **result})


@api_view(["POST"])
def send_alerts(request):
    """Check for critical alerts and send notifications via configured channels."""
    bot_id = request.data.get("bot_id")

    if not bot_id:
        return Response(
            {"error": "bot_id is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    result = check_and_send_alerts(bot_id)

    return Response({"status": "success", **result})


# ──────────────────────────────────────────────────────────────
# Sprint 6: Data Integration Layer (Connectors)
# ──────────────────────────────────────────────────────────────

@api_view(['GET', 'POST'])
def data_sources(request):
    """
    GET  /api/connectors/          — list all data sources for a bot_id
    POST /api/connectors/          — register a new data source connection
    """
    bot_id = get_verified_bot_id(request, request.query_params.get('bot_id') or request.data.get('bot_id'))
    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'GET':
        sources = DataSource.objects.filter(bot_id=bot_id).values(
            'id', 'source_type', 'display_name', 'status', 'last_synced_at', 'created_at'
        )
        return Response(list(sources))

    # POST: create a new connector entry
    source_type = request.data.get('source_type')
    display_name = request.data.get('display_name', source_type)
    config = request.data.get('config', {})

    VALID_TYPES = ['tally', 'razorpay', 'google_sheets', 'zoho']
    if source_type not in VALID_TYPES:
        return Response(
            {'error': f'Invalid source_type. Choose from: {VALID_TYPES}'},
            status=status.HTTP_400_BAD_REQUEST
        )

    ds = DataSource.objects.create(
        bot_id=bot_id,
        source_type=source_type,
        display_name=display_name,
        config=config,
    )
    return Response({
        'id': ds.id,
        'source_type': ds.source_type,
        'display_name': ds.display_name,
        'status': ds.status,
        'message': 'Data source registered successfully.',
    }, status=status.HTTP_201_CREATED)


@api_view(['DELETE'])
def delete_data_source(request, source_id):
    """DELETE /api/connectors/<id>/ — remove a registered connector."""
    bot_id = get_verified_bot_id(request, request.query_params.get('bot_id') or request.data.get('bot_id'))
    try:
        ds = DataSource.objects.get(id=source_id, bot_id=bot_id)
        ds.delete()
        return Response({'message': 'Connector deleted.'}, status=status.HTTP_200_OK)
    except DataSource.DoesNotExist:
        return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
def trigger_sync(request, source_id):
    """
    POST /api/connectors/<id>/sync/ — manually trigger a data sync for a connector.
    Imports transactions from the external source into our database.
    """
    from .connectors import get_connector

    bot_id = get_verified_bot_id(request, request.data.get('bot_id')) or request.query_params.get('bot_id')
    try:
        ds = DataSource.objects.get(id=source_id, bot_id=bot_id)
    except DataSource.DoesNotExist:
        return Response({'error': 'Data source not found.'}, status=status.HTTP_404_NOT_FOUND)

    try:
        connector = get_connector(ds)
        log = connector.sync()
        return Response({
            'status': log.status,
            'records_fetched': log.records_fetched,
            'records_inserted': log.records_inserted,
            'currency': log.currency_used,
            'exchange_rate': str(log.exchange_rate),
            'error': log.error_message,
            'synced_at': log.synced_at.isoformat(),
        })
    except ValueError as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': f'Sync failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ──────────────────────────────────────────────────────────────
# Sprint 7: Advanced Budgeting & Forecasting
# ──────────────────────────────────────────────────────────────

class BudgetListCreateView(generics.ListCreateAPIView):
    """
    GET /api/budgets/ - List active budgets for bot
    POST /api/budgets/ - Create or update a budget (upsert logic to handle versions)
    """
    serializer_class = BudgetSerializer

    def get_queryset(self):
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get('bot_id'))
        month = self.request.query_params.get('month_year')
        qs = Budget.objects.filter(is_active=True)
        if bot_id:
            qs = qs.filter(bot_id=bot_id)
        if month:
            qs = qs.filter(month_year=month)
        return qs.order_by('month_year', 'category')

    def create(self, request, *args, **kwargs):
        # We want to do "upsert": if a budget for this bot, category, month exists, 
        # deactivate it and create a v+1
        bot_id = get_verified_bot_id(request, request.data.get('bot_id'))
        category = request.data.get('category')
        month_year = request.data.get('month_year')
        allocated_amount = request.data.get('allocated_amount')

        if not all([bot_id, category, month_year, allocated_amount is not None]):
            return Response({"error": "Missing required fields"}, status=status.HTTP_400_BAD_REQUEST)

        # Find existing active budget
        existing = Budget.objects.filter(bot_id=bot_id, category=category, month_year=month_year, is_active=True).first()
        version = 1
        if existing:
            version = existing.version + 1
            existing.is_active = False
            existing.save()
            
        new_budget = Budget.objects.create(
            bot_id=bot_id,
            category=category,
            month_year=month_year,
            allocated_amount=allocated_amount,
            version=version,
            is_active=True
        )
        
        return Response(BudgetSerializer(new_budget).data, status=status.HTTP_201_CREATED)

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_budget(request):
    """
    POST /api/budgets/upload/
    Uploads an Excel file containing headers: [category, allocated_amount, month_year]
    """
    bot_id = get_verified_bot_id(request, request.data.get('bot_id'))
    file_obj = request.FILES.get('file')

    if not bot_id or not file_obj:
        return Response({"error": "bot_id and file are required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        import pandas as pd
        df = pd.read_excel(file_obj)
        # Process expected columns
        required_cols = {'category', 'allocated_amount', 'month_year'}
        if not required_cols.issubset(set(df.columns)):
            return Response({"error": f"Excel file must contain columns: {required_cols}"}, status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        for _, row in df.iterrows():
            category = str(row['category']).strip()
            amount = float(row['allocated_amount'])
            month = str(row['month_year']).strip()

            existing = Budget.objects.filter(bot_id=bot_id, category=category, month_year=month, is_active=True).first()
            version = 1
            if existing:
                version = existing.version + 1
                existing.is_active = False
                existing.save()
                
            Budget.objects.create(
                bot_id=bot_id,
                category=category,
                month_year=month,
                allocated_amount=amount,
                version=version,
                is_active=True
            )
            created_count += 1

        return Response({"message": f"Successfully parsed and saved {created_count} budget entries."}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def variance_analysis(request):
    """
    GET /api/budgets/variance/?bot_id=xyz&month_year=2026-03
    """
    bot_id = request.query_params.get('bot_id')
    month = request.query_params.get('month_year')
    if not bot_id or not month:
        return Response({"error": "bot_id and month_year required"}, status=status.HTTP_400_BAD_REQUEST)

    result = calculate_variance(bot_id, month)
    if "error" in result:
        return Response(result, status=status.HTTP_200_OK)
        
    return Response(result)

@api_view(['GET'])
def monte_carlo_simulation(request):
    """
    GET /api/forecast/monte-carlo/?bot_id=xyz
    """
    bot_id = request.query_params.get('bot_id')
    if not bot_id:
        return Response({"error": "bot_id is required"}, status=status.HTTP_400_BAD_REQUEST)

    result = run_monte_carlo_simulation(bot_id)
    if "error" in result:
        return Response(result, status=status.HTTP_200_OK)
        
    return Response(result)

# ──────────────────────────────────────────────────────────────
# Sprint 8: Invoice & AP Automation
# ──────────────────────────────────────────────────────────────

class PurchaseOrderListCreateView(generics.ListCreateAPIView):
    """
    GET /api/purchase-orders/?bot_id=123
    POST /api/purchase-orders/
    """
    serializer_class = PurchaseOrderSerializer

    def get_queryset(self):
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get('bot_id'))
        qs = PurchaseOrder.objects.all()
        if bot_id:
            qs = qs.filter(bot_id=bot_id)
        return qs.order_by('-created_at')

class InvoiceListView(generics.ListAPIView):
    """
    GET /api/invoices/?bot_id=123
    """
    serializer_class = InvoiceSerializer

    def get_queryset(self):
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get('bot_id'))
        qs = Invoice.objects.all()
        if bot_id:
            qs = qs.filter(bot_id=bot_id)
        return qs.order_by('-uploaded_at')

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_invoice(request):
    """
    POST /api/invoices/upload/
    Uploads an image or PDF of an invoice for Gemini Vision extraction and fraud reasoning.
    """
    from .services.invoice_processor import process_invoice_document
    import os

    bot_id = get_verified_bot_id(request, request.data.get('bot_id'))
    file_obj = request.FILES.get('file')

    if not bot_id or not file_obj:
        return Response({"error": "bot_id and file are required"}, status=status.HTTP_400_BAD_REQUEST)

    # Save temp file
    upload_dir = os.path.join('media', 'invoices')
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, file_obj.name)
    
    with open(file_path, 'wb+') as destination:
        for chunk in file_obj.chunks():
            destination.write(chunk)

    # Convert to bytes for Gemini
    with open(file_path, 'rb') as f:
        file_bytes = f.read()
    
    mime_type = "application/pdf" if file_obj.name.endswith('.pdf') else "image/jpeg"
    if file_obj.name.endswith('.png'):
        mime_type = "image/png"

    # Send to multimodal AI
    result = process_invoice_document(bot_id, file_path, mime_type, file_bytes)
    
    if "error" in result:
        return Response(result, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response(result, status=status.HTTP_201_CREATED)

@api_view(['PATCH'])
def update_invoice_status(request, invoice_id):
    """
    PATCH /api/invoices/<id>/status/
    Updates the status of an Invoice (e.g., Pending -> Approved or Rejected).
    """
    try:
        invoice = Invoice.objects.get(id=invoice_id)
        
        # Verify Bot ID Authorization
        bot_id = get_verified_bot_id(request, request.data.get('bot_id')) or request.query_params.get('bot_id')
        if not bot_id or invoice.bot_id != bot_id:
            return Response({'error': 'Unauthorized Workspace'}, status=status.HTTP_403_FORBIDDEN)

        new_status = request.data.get('status')
        if new_status not in ['pending_approval', 'approved', 'rejected', 'paid']:
            return Response({'error': 'Invalid status provided.'}, status=status.HTTP_400_BAD_REQUEST)

        invoice.status = new_status
        invoice.save()
        
        return Response({'success': True, 'status': invoice.status})
        
    except Invoice.DoesNotExist:
        return Response({'error': 'Invoice not found.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Error updating invoice: {str(e)}")
        return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(['GET'])
def notifications_count(request):
    """
    GET /api/notifications/count/?bot_id=123
    Returns the number of unseen announcements since last_seen_at.
    """
    bot_id = request.query_params.get('bot_id')
    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    meta, _ = NotificationMeta.objects.get_or_create(bot_id=bot_id)
    last_seen = meta.last_seen_at

    # Count how many of each notification type were created after 'last_seen'
    unseen_anomalies = AnomalyLog.objects.filter(bot_id=bot_id, detected_at__gt=last_seen).count()
    unseen_recs = Recommendation.objects.filter(bot_id=bot_id, created_at__gt=last_seen).count()
    unseen_uploads = UploadedFile.objects.filter(bot_id=bot_id, uploaded_at__gt=last_seen).count()

    total_unseen = unseen_anomalies + unseen_recs + unseen_uploads

    return Response({'count': total_unseen})


@api_view(['GET'])
def notifications_list(request):
    """
    GET /api/notifications/?bot_id=123
    Aggregates news from Anomalies, Recommendations, and File Uploads.
    Also updates last_seen_at for the bot_id.
    """
    bot_id = request.query_params.get('bot_id')
    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    # Mark as seen
    meta, _ = NotificationMeta.objects.get_or_create(bot_id=bot_id)
    meta.last_seen_at = timezone.now()
    meta.save()

    # 1. Fetch Anomalies
    anomalies = AnomalyLog.objects.filter(bot_id=bot_id).order_by('-detected_at')[:10]
    
    # 2. Fetch Recommendations
    recommendations = Recommendation.objects.filter(bot_id=bot_id).order_by('-created_at')[:10]

    # 3. Fetch Recent Uploads
    uploads = UploadedFile.objects.filter(bot_id=bot_id).order_by('-uploaded_at')[:10]

    # Format into a unified structure
    unified = []

    for a in anomalies:
        unified.append({
            'id': f"anomaly_{a.id}",
            'title': f"{a.severity.upper()} Spend Alert: {a.category}",
            'description': a.description,
            'time': a.detected_at.isoformat(),
            'type': 'fraud' if a.severity in ['high', 'critical'] else 'warning',
        })

    for r in recommendations:
        unified.append({
            'id': f"rec_{r.id}",
            'title': f"AI CFO Advice: {r.title}",
            'description': r.detail,
            'time': r.created_at.isoformat(),
            'type': 'info' if r.priority == 'info' else 'action',
        })

    for u in uploads:
        unified.append({
            'id': f"upload_{u.id}",
            'title': f"File Processed: {u.original_filename}",
            'description': f"Successfully parsed {u.row_count} records." if u.status == 'completed' else f"Failed: {u.error_message}",
            'time': u.uploaded_at.isoformat(),
            'type': 'success' if u.status == 'completed' else 'error',
        })

    # Sort descending by time
    unified.sort(key=lambda x: x['time'], reverse=True)

    return Response(unified[:20])


# =====================================================
# Sprint 10: Financial Reporting & Dashboards
# =====================================================

from django.http import HttpResponse

@api_view(['GET'])
def report_pnl(request):
    from api.services.reporting_logic import generate_pnl
    bot_id = request.query_params.get('bot_id')
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')

    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    data = generate_pnl(bot_id, start_date, end_date)
    return Response(data)

@api_view(['GET'])
def report_cashflow(request):
    from api.services.reporting_logic import generate_cashflow
    bot_id = request.query_params.get('bot_id')
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')

    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    data = generate_cashflow(bot_id, start_date, end_date)
    return Response(data)

@api_view(['GET'])
def report_balancesheet(request):
    from api.services.reporting_logic import generate_balancesheet
    bot_id = request.query_params.get('bot_id')
    target_date = request.query_params.get('target_date')

    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    data = generate_balancesheet(bot_id, target_date)
    return Response(data)

@api_view(['GET'])
def report_export_excel(request):
    """
    Generates an Excel document for the requested report type using pandas.
    """
    from api.services.reporting_logic import generate_pnl, generate_cashflow, generate_balancesheet
    import pandas as pd
    bot_id = request.query_params.get('bot_id')
    report_type = request.query_params.get('type')  # pnl, cashflow, balancesheet
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    
    if not bot_id or not report_type:
        return Response({'error': 'bot_id and type are required'}, status=status.HTTP_400_BAD_REQUEST)

    data = None
    if report_type == 'pnl':
        data = generate_pnl(bot_id, start_date, end_date)
        # Flatten dictionary to list of dicts for pandas
        rows = []
        for section in ['Revenue', 'COGS', 'OPEX']:
            for item, amt in data.get(section, {}).get('items', {}).items():
                rows.append({'Section': section, 'Account': item, 'Amount': amt})
        df = pd.DataFrame(rows)

    elif report_type == 'cashflow':
        data = generate_cashflow(bot_id, start_date, end_date)
        rows = []
        for section in ['OperatingActivities', 'InvestingActivities', 'FinancingActivities']:
            rows.append({'Section': section, 'Inflows': data[section]['inflows'], 'Outflows': data[section]['outflows'], 'Net': data[section]['net_cash']})
        df = pd.DataFrame(rows)

    elif report_type == 'balancesheet':
        data = generate_balancesheet(bot_id, end_date or None)
        rows = []
        for section in ['Assets', 'Liabilities', 'Equity']:
            for item, amt in data.get(section, {}).get('items', {}).items():
                rows.append({'Section': section, 'Account': item, 'Amount': amt})
        df = pd.DataFrame(rows)
    else:
        return Response({'error': 'invalid report type'}, status=status.HTTP_400_BAD_REQUEST)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=financial_report_{report_type}_{bot_id}.xlsx'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Report')
        workbook = writer.book
        worksheet = writer.sheets['Report']

        # Format Headers
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E2637")
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Freeze top row
        worksheet.freeze_panes = "A2"

        # Apply Currency formatting to amount columns
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                col_letter = worksheet.cell(row=1, column=col).column_letter
                header_name = worksheet.cell(row=1, column=col).value
                if header_name in ['Amount', 'Inflows', 'Outflows', 'Net']:
                    worksheet[f"{col_letter}{row}"].number_format = '"$"#,##0.00'

        # Auto-adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

    return response

# =====================================================
# Sprint 11: Enterprise RBAC & Team Management
# =====================================================
from .models import UserFeaturePermission
from .decorators import require_feature

@api_view(['GET'])
def get_team_permissions(request):
    """
    Returns the granular feature permissions for all users in the organization.
    Must be called by an org:admin.
    """
    bot_id = request.query_params.get('bot_id')
    org_role = request.headers.get('X-Org-Role')
    
    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
    if org_role != 'org:admin':
        return Response({'error': 'Only Admins can view team permissions'}, status=status.HTTP_403_FORBIDDEN)

    permissions = UserFeaturePermission.objects.filter(bot_id=bot_id)
    data = []
    for p in permissions:
        data.append({
            'user_id': p.user_id,
            'user_email': p.user_email,
            'can_view_reports': p.can_view_reports,
            'can_upload_data': p.can_upload_data,
            'can_manage_budgets': p.can_manage_budgets,
            'can_manage_ap': p.can_manage_ap,
            'can_run_simulations': p.can_run_simulations,
        })
    return Response(data)

@api_view(['POST'])
def update_team_permission(request):
    """
    Overwrites a specific user's granular feature boolean.
    """
    bot_id = get_verified_bot_id(request, request.data.get('bot_id'))
    org_role = request.headers.get('X-Org-Role')
    target_user_id = request.data.get('user_id')
    feature_name = request.data.get('feature')  # e.g., 'can_view_reports'
    new_value = request.data.get('value')       # bool
    
    if org_role != 'org:admin':
        return Response({'error': 'Only Admins can modify team permissions'}, status=status.HTTP_403_FORBIDDEN)

    try:
        perm = UserFeaturePermission.objects.get(bot_id=bot_id, user_id=target_user_id)
        if hasattr(perm, feature_name):
            setattr(perm, feature_name, new_value)
            perm.save()
            return Response({'success': True, 'message': f'Updated {feature_name} to {new_value}'})
        else:
            return Response({'error': 'Invalid feature name'}, status=status.HTTP_400_BAD_REQUEST)
    except UserFeaturePermission.DoesNotExist:
        # If the user doesn't exist in our DB yet, create them with defaults
        user_email = request.data.get('user_email', 'unknown@email.com')
        perm = UserFeaturePermission(bot_id=bot_id, user_id=target_user_id, user_email=user_email)
        if hasattr(perm, feature_name):
            setattr(perm, feature_name, new_value)
        perm.save()
        return Response({'success': True, 'message': f'Created rules and set {feature_name} to {new_value}'})

# =====================================================
# Sprint 11: Phase 2 — Collaborative Anomaly Workflows
# =====================================================
from .models import AnomalyComment
from .services.audit_service import log_event
import json


@api_view(['PATCH'])
def update_anomaly_status(request, anomaly_id):
    """
    PATCH /api/anomalies/<id>/status/
    Updates resolution status and/or assigned_to for an Anomaly.
    """
    try:
        bot_id = get_verified_bot_id(request, request.data.get('bot_id')) or request.query_params.get('bot_id')
        anomaly = AnomalyLog.objects.get(id=anomaly_id, bot_id=bot_id)

        new_status = request.data.get('status')
        assigned_to = request.data.get('assigned_to')

        if new_status and new_status in ['open', 'in-review', 'resolved']:
            anomaly.status = new_status
            if new_status == 'resolved':
                anomaly.is_resolved = True
        if assigned_to is not None:
            anomaly.assigned_to = assigned_to

        anomaly.save()
        log_event(request, 'ANOMALY_STATUS_UPDATE', 'Anomaly', anomaly_id,
                  json.dumps({'newStatus': new_status, 'assignedTo': assigned_to}))

        return Response({'success': True, 'status': anomaly.status, 'assigned_to': anomaly.assigned_to})

    except AnomalyLog.DoesNotExist:
        return Response({'error': 'Anomaly not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET', 'POST'])
def anomaly_comments(request, anomaly_id):
    """
    GET  /api/anomalies/<id>/comments/ — fetch all comments
    POST /api/anomalies/<id>/comments/ — add a new comment with @mention detection
    """
    bot_id = get_verified_bot_id(request, request.query_params.get('bot_id') or request.data.get('bot_id'))

    if request.method == 'GET':
        comments = AnomalyComment.objects.filter(anomaly_id=anomaly_id, bot_id=bot_id).order_by('created_at')
        data = [
            {'id': c.id, 'user_id': c.user_id, 'user_email': c.user_email,
             'text': c.text, 'created_at': c.created_at.isoformat()}
            for c in comments
        ]
        return Response(data)

    elif request.method == 'POST':
        user_id = request.headers.get('X-User-Id', 'anonymous')
        user_email = request.headers.get('X-User-Email', request.data.get('user_email', 'user@cfol.ai'))
        text = request.data.get('text', '').strip()

        if not text:
            return Response({'error': 'Comment text is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            anomaly = AnomalyLog.objects.get(id=anomaly_id, bot_id=bot_id)
        except AnomalyLog.DoesNotExist:
            return Response({'error': 'Anomaly not found'}, status=status.HTTP_404_NOT_FOUND)

        comment = AnomalyComment.objects.create(
            bot_id=bot_id, anomaly=anomaly,
            user_id=user_id, user_email=user_email, text=text,
        )

        # @mention email trigger
        import re
        mentions = re.findall(r'@([\w.+-]+@[\w-]+\.[a-zA-Z]+)', text)
        if mentions:
            try:
                from django.core.mail import send_mail
                for email in mentions:
                    send_mail(
                        subject=f'[CFOlytics] {user_email} mentioned you in an anomaly',
                        message=(
                            f'Hi,\n\n{user_email} mentioned you on a {anomaly.severity.upper()} anomaly:\n\n'
                            f'Anomaly: "{anomaly.description[:100]}"\n\n'
                            f'Comment: "{text}"\n\nLog in to CFOlytics to respond.\n\nCFOlytics'
                        ),
                        from_email='noreply@cfol.ai',
                        recipient_list=[email],
                        fail_silently=True,
                    )
            except Exception:
                pass

        log_event(request, 'ANOMALY_COMMENT_ADDED', 'Anomaly', anomaly_id, text[:200])

        return Response({
            'id': comment.id, 'user_email': comment.user_email,
            'text': comment.text, 'created_at': comment.created_at.isoformat(),
        }, status=status.HTTP_201_CREATED)


# =====================================================
# Sprint 11: Phase 3 — Immutable Audit Trail
# =====================================================
from .models import AuditEvent
import csv
from django.http import HttpResponse as DjangoHttpResponse


@api_view(['GET'])
def audit_trail(request):
    """
    GET /api/audit/?bot_id=xxx
    Returns paginated audit log. Admin-only.
    """
    bot_id = request.query_params.get('bot_id')
    org_role = request.headers.get('X-Org-Role')

    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)
    if org_role != 'org:admin':
        return Response({'error': 'Only Admins can view audit logs'}, status=status.HTTP_403_FORBIDDEN)

    page = int(request.query_params.get('page', 1))
    per_page = 50
    offset = (page - 1) * per_page

    events = AuditEvent.objects.filter(bot_id=bot_id).order_by('-timestamp')[offset: offset + per_page]
    total = AuditEvent.objects.filter(bot_id=bot_id).count()

    data = [
        {
            'id': e.id,
            'user_email': e.user_email,
            'action': e.action,
            'resource_type': e.resource_type,
            'resource_id': e.resource_id,
            'details': e.details,
            'ip_address': str(e.ip_address) if e.ip_address else None,
            'timestamp': e.timestamp.isoformat(),
        }
        for e in events
    ]
    return Response({'results': data, 'total': total, 'page': page})


@api_view(['GET'])
def audit_export_csv(request):
    """
    GET /api/audit/export/?bot_id=xxx
    Streams full audit trail as CSV. Admin-only.
    """
    bot_id = request.query_params.get('bot_id')
    org_role = request.headers.get('X-Org-Role')

    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)
    if org_role != 'org:admin':
        return Response({'error': 'Only Admins can export audit logs'}, status=status.HTTP_403_FORBIDDEN)

    events = AuditEvent.objects.filter(bot_id=bot_id).order_by('-timestamp')
    response = DjangoHttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="audit_trail_{bot_id}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Timestamp', 'User Email', 'Action', 'Resource Type', 'Resource ID', 'Details', 'IP Address'])
    for e in events:
        writer.writerow([
            e.timestamp.strftime('%Y-%m-%d %H:%M:%S'), e.user_email, e.action,
            e.resource_type, e.resource_id, e.details, e.ip_address or '',
        ])
    return response

from .services.proactive_engine import generate_morning_brief

@api_view(['GET'])
def daily_briefing(request):
    """
    GET /api/briefing/?bot_id=123
    Returns the daily AI-generated executive briefing.
    """
    bot_id = request.query_params.get('bot_id')
    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    brief = generate_morning_brief(bot_id)
    return Response(brief)

# ──────────────────────────────────────────────
# Sprint 15: Workspace Architecture & Org Chat
# ──────────────────────────────────────────────

class WorkspaceListCreateView(generics.ListCreateAPIView):
    """
    GET  /api/workspaces/?org_id=xxx&status=active&q=searchterm
    POST /api/workspaces/  - create workspace (auto-hashes secure_key)
    """
    serializer_class = WorkspaceSerializer

    def get_queryset(self):
        org_id = self.request.query_params.get("org_id")
        ws_status = self.request.query_params.get("status")
        q = self.request.query_params.get("q", "")
        qs = Workspace.objects.all().order_by('-created_at')
        if org_id:
            qs = qs.filter(org_id=org_id)
        if ws_status:
            qs = qs.filter(status=ws_status)
        else:
            # Hide closed workspaces from main list by default
            qs = qs.exclude(status=Workspace.STATUS_CLOSED)
        if q:
            qs = qs.filter(name__icontains=q)
        return qs


class WorkspaceRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    """Soft-deletes by setting status=closed instead of hard-deleting."""
    queryset = Workspace.objects.all()
    serializer_class = WorkspaceSerializer

    def destroy(self, request, *args, **kwargs):
        workspace = self.get_object()
        workspace.status = Workspace.STATUS_CLOSED
        workspace.save()
        return Response({'status': 'closed'}, status=status.HTTP_200_OK)


@api_view(['PATCH'])
def workspace_set_status(request, pk):
    """
    PATCH /api/workspaces/<id>/status/
    Body: {"status": "active"|"paused"|"closed"}
    Allows instant lifecycle transitions without going through full PATCH.
    """
    try:
        workspace = Workspace.objects.get(pk=pk)
    except Workspace.DoesNotExist:
        return Response({'error': 'Workspace not found'}, status=status.HTTP_404_NOT_FOUND)

    new_status = request.data.get('status')
    valid = [Workspace.STATUS_ACTIVE, Workspace.STATUS_PAUSED, Workspace.STATUS_CLOSED]
    if new_status not in valid:
        return Response({'error': 'Invalid status value'}, status=status.HTTP_400_BAD_REQUEST)

    workspace.status = new_status
    workspace.save()
    return Response({'id': workspace.id, 'name': workspace.name, 'status': workspace.status})

class GoalTargetListCreateView(generics.ListCreateAPIView):
    """
    List/Create Goals for KPI Radial tracking per workspace.
    """
    serializer_class = GoalTargetSerializer

    def get_queryset(self):
        bot_id = get_verified_bot_id(self.request, self.request.query_params.get("bot_id"))
        workspace_id = self.request.query_params.get("workspace_id")
        qs = GoalTarget.objects.all().order_by('-created_at')
        if bot_id:
            qs = qs.filter(bot_id=bot_id)
        if workspace_id:
            qs = qs.filter(workspace_id=workspace_id)
        return qs

class OrgChatMessageListCreateView(generics.ListCreateAPIView):
    """
    Simple real-time organization chat messaging polling endpoint.
    """
    serializer_class = OrgChatMessageSerializer

    def get_queryset(self):
        org_id = self.request.query_params.get("org_id")
        limit = int(self.request.query_params.get("limit", 50))
        qs = OrgChatMessage.objects.all().order_by('-created_at')
        if org_id:
            qs = qs.filter(org_id=org_id)
        return qs[:limit]


# =====================================================
# Sprint 18: Invoice Workspace (Contextual Chat Threads)
# =====================================================

@api_view(['GET', 'POST'])
def invoice_threads(request, invoice_id):
    """
    GET  /api/invoices/<id>/threads/  — list all thread messages on an invoice
    POST /api/invoices/<id>/threads/  — post a new message to the invoice workspace
    Body: { bot_id, user_id, user_email, user_name, message }
    Mentions: detected automatically from @email patterns in the message text.
    """
    import re
    try:
        invoice = Invoice.objects.get(pk=invoice_id)
    except Invoice.DoesNotExist:
        return Response({'error': 'Invoice not found.'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        threads = InvoiceThread.objects.filter(invoice=invoice).order_by('created_at')
        return Response([{
            'id': t.id,
            'user_id': t.user_id,
            'user_email': t.user_email,
            'user_name': t.user_name or t.user_email,
            'message': t.message,
            'mentions': t.mentions,
            'created_at': t.created_at.isoformat(),
        } for t in threads])

    # POST: create a new thread message
    bot_id = request.data.get('bot_id', '')
    user_id = request.data.get('user_id', '')
    user_email = request.data.get('user_email', '')
    user_name = request.data.get('user_name', '')
    message = request.data.get('message', '').strip()

    if not message or not user_email:
        return Response({'error': 'message and user_email are required.'}, status=status.HTTP_400_BAD_REQUEST)

    # Auto-detect @mentions (patterns like @someone@company.com)
    mention_pattern = r'@([\w.+-]+@[\w-]+\.[\w.]+)'
    mentions = re.findall(mention_pattern, message)

    thread = InvoiceThread.objects.create(
        bot_id=bot_id,
        invoice=invoice,
        user_id=user_id,
        user_email=user_email,
        user_name=user_name,
        message=message,
        mentions=mentions,
    )

    # Fire notification to in-app notification system for @mentions
    if mentions:
        for mentioned_email in mentions:
            AuditEvent.objects.create(
                bot_id=bot_id,
                user_id=user_id,
                user_email=user_email,
                action='INVOICE_MENTION',
                resource_type='Invoice',
                resource_id=str(invoice_id),
                details=f'{user_name or user_email} mentioned {mentioned_email} on Invoice #{invoice_id}: "{message[:100]}"'
            )

    return Response({
        'id': thread.id,
        'user_email': thread.user_email,
        'user_name': thread.user_name,
        'message': thread.message,
        'mentions': thread.mentions,
        'created_at': thread.created_at.isoformat(),
    }, status=status.HTTP_201_CREATED)


# =====================================================
# Sprint 18: VicInbox — Mailgun Email Webhook Handler
# =====================================================

@api_view(['POST'])
def mailgun_inbound_webhook(request):
    """
    POST /api/ap/webhooks/mailgun-inbound/

    Receives inbound email payloads from Mailgun's Inbound Routing.
    Extracts PDF/image attachments and runs them through the Gemini
    invoice processor automatically (Autopilot).

    For local testing, you can POST a JSON body mimicking Mailgun's format:
    {
      "sender":     "vendor@example.com",
      "recipient":  "invoices.org123@cfolytics.com",
      "subject":    "Invoice #INV-2026-042",
      "bot_id":     "org_...",          // only needed for mock/testing
      "attachment-count": "1"           // Mailgun includes this
    }
    Real Mailgun webhooks include 'attachment-1' as a multipart file.
    """
    from .services.invoice_processor import process_invoice_document
    import os, base64, hmac, hashlib

    # Mailgun Signature Verification
    signing_key = os.environ.get('MAILGUN_SIGNING_KEY')
    if signing_key:
        timestamp = request.data.get('timestamp', '')
        token = request.data.get('token', '')
        signature = request.data.get('signature', '')
        
        expected_sig = hmac.new(
            key=signing_key.encode('utf-8'),
            msg=f"{timestamp}{token}".encode('utf-8'),
            digestmod=hashlib.sha256
        ).hexdigest()
        
        if not hmac.compare_digest(signature, expected_sig):
            return Response({'error': 'Invalid webhook signature'}, status=status.HTTP_403_FORBIDDEN)

    sender = request.data.get('sender') or request.data.get('from', 'unknown@sender.com')
    recipient = request.data.get('recipient', '')
    subject = request.data.get('subject', '')

    # Derive bot_id from recipient address (e.g. invoices.org_abc123@cfolytics.com -> org_abc123)
    # For mock testing, accept explicit bot_id in payload
    bot_id = request.data.get('bot_id', '')
    if not bot_id and recipient:
        # Parse: invoices.<bot_id>@cfolytics.com
        local_part = recipient.split('@')[0]  # e.g. "invoices.org_abc"
        parts = local_part.split('.', 1)
        if len(parts) == 2:
            bot_id = parts[1]

    # Log the inbound email immediately
    log = InboundEmailLog.objects.create(
        bot_id=bot_id,
        sender_email=sender,
        subject=subject,
        recipient=recipient,
        raw_payload={
            'sender': sender,
            'recipient': recipient,
            'subject': subject,
            'attachment_count': request.data.get('attachment-count', 0),
        },
        status='processing',
    )

    # Check for file attachments
    attachment = request.FILES.get('attachment-1') or request.FILES.get('file')

    if not attachment:
        log.status = 'no_attachment'
        log.error_message = 'Email received but no PDF/image attachment found.'
        log.save()
        return Response({
            'status': 'no_attachment',
            'message': 'Email received but no invoice attachment found.',
            'log_id': log.id,
        }, status=status.HTTP_200_OK)

    # Save the attachment temporarily
    upload_dir = os.path.join('media', 'email_invoices')
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, f'email_{log.id}_{attachment.name}')

    with open(file_path, 'wb+') as dest:
        for chunk in attachment.chunks():
            dest.write(chunk)

    with open(file_path, 'rb') as f:
        file_bytes = f.read()

    mime_type = 'application/pdf'
    if attachment.name.lower().endswith('.png'):
        mime_type = 'image/png'
    elif attachment.name.lower().endswith(('.jpg', '.jpeg')):
        mime_type = 'image/jpeg'

    if not bot_id:
        log.status = 'failed'
        log.error_message = 'Could not determine bot_id from recipient address. Send to invoices.<bot_id>@cfolytics.com'
        log.save()
        return Response({'error': 'Cannot determine workspace. Use invoices.<bot_id>@cfolytics.com as recipient.'}, status=status.HTTP_400_BAD_REQUEST)

    # Process through Gemini Vision invoice processor
    result = process_invoice_document(bot_id, file_path, mime_type, file_bytes)

    if 'error' in result:
        log.status = 'failed'
        log.error_message = result['error']
        log.save()
        return Response({'status': 'failed', 'error': result['error'], 'log_id': log.id}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Link the created invoice back to the log
    try:
        invoice = Invoice.objects.get(pk=result['invoice_id'])
        log.invoice = invoice
    except Invoice.DoesNotExist:
        pass

    log.status = 'processed'
    log.save()

    return Response({
        'status': 'processed',
        'log_id': log.id,
        'invoice_id': result.get('invoice_id'),
        'vendor_name': result.get('vendor_name'),
        'total_amount': result.get('total_amount'),
        'autopilot_approved': result.get('autopilot_approved', False),
        'fraud_score': result.get('fraud_score'),
        'invoice_status': result.get('status'),
        'message': f'Invoice from {sender} processed successfully via email ingestion.',
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
def email_inbox_logs(request):
    """
    GET /api/ap/email-logs/?bot_id=xxx
    Lists inbound email processing history for a workspace.
    """
    bot_id = request.query_params.get('bot_id')
    if not bot_id:
        return Response({'error': 'bot_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    logs = InboundEmailLog.objects.filter(bot_id=bot_id).order_by('-received_at')[:50]
    return Response([{
        'id': l.id,
        'sender_email': l.sender_email,
        'subject': l.subject,
        'status': l.status,
        'invoice_id': l.invoice.id if l.invoice else None,
        'error_message': l.error_message,
        'received_at': l.received_at.isoformat(),
    } for l in logs])


# =====================================================
# Sprint 18 Part B: Ad-Hoc Data Query Agent (NL2SQL)
# =====================================================

@api_view(['POST'])
def nl_query(request):
    """
    POST /api/query/
    Body: { "bot_id": "...", "question": "Show me top 10 expenses" }

    Converts natural-language financial questions into safe Django ORM queries.
    Returns: { success, intent, summary, columns, rows }
    """
    from .services.nl2sql_service import run_nl_query

    bot_id = get_verified_bot_id(request, request.data.get('bot_id'))
    question = request.data.get('question', '').strip()

    if not bot_id or not question:
        return Response({'error': 'bot_id and question are required.'}, status=status.HTTP_400_BAD_REQUEST)

    result = run_nl_query(bot_id, question)
    return Response(result)


# =====================================================
# Sprint 18 Part B: Generative Budget Planner
# =====================================================

@api_view(['POST'])
def generate_budget(request):
    """
    POST /api/budget/generate/
    Body: {
        "bot_id": "...",
        "target_month": "2026-06",
        "growth_assumption": 15,
        "instructions": "Cut travel by 20%",
        "apply_to_db": false
    }

    Uses Gemini to synthesize a full AI budget plan modeled on historical data.
    Returns: { success, budget_items, ai_rationale, total_budget }
    """
    from .services.budget_generator_service import generate_ai_budget

    bot_id = get_verified_bot_id(request, request.data.get('bot_id'))
    if not bot_id:
        return Response({'error': 'bot_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    params = {
        'target_month': request.data.get('target_month'),
        'growth_assumption': request.data.get('growth_assumption', 10),
        'instructions': request.data.get('instructions', ''),
        'apply_to_db': request.data.get('apply_to_db', False),
    }

    result = generate_ai_budget(bot_id, params)

    if not result.get('success'):
        return Response(result, status=status.HTTP_400_BAD_REQUEST)

    return Response(result, status=status.HTTP_200_OK)


# =====================================================
# Sprint 18 Part B: Custom KPI Builder
# =====================================================

@api_view(['GET', 'POST'])
def custom_kpis(request):
    """
    GET  /api/kpi-builder/?bot_id=xxx      — list all custom KPIs
    POST /api/kpi-builder/                 — create a new custom KPI
    Body: { bot_id, name, description, formula, unit, icon, color }
    """
    from .models import CustomKPI

    bot_id = get_verified_bot_id(request, request.query_params.get('bot_id') or request.data.get('bot_id'))
    if not bot_id:
        return Response({'error': 'bot_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'GET':
        kpis = CustomKPI.objects.filter(bot_id=bot_id, is_active=True)
        return Response([{
            'id': k.id,
            'name': k.name,
            'description': k.description,
            'formula': k.formula,
            'unit': k.unit,
            'icon': k.icon,
            'color': k.color,
            'sort_order': k.sort_order,
            'created_at': k.created_at.isoformat(),
        } for k in kpis])

    # POST — create
    name = request.data.get('name', '').strip()
    formula = request.data.get('formula', '').strip()
    if not name or not formula:
        return Response({'error': 'name and formula are required.'}, status=status.HTTP_400_BAD_REQUEST)

    kpi = CustomKPI.objects.create(
        bot_id=bot_id,
        name=name,
        description=request.data.get('description', ''),
        formula=formula,
        unit=request.data.get('unit', '$'),
        icon=request.data.get('icon', '📊'),
        color=request.data.get('color', 'indigo'),
    )

    return Response({
        'id': kpi.id,
        'name': kpi.name,
        'formula': kpi.formula,
        'unit': kpi.unit,
        'icon': kpi.icon,
        'color': kpi.color,
        'message': f'Custom KPI "{kpi.name}" created successfully.',
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
def evaluate_custom_kpi(request, kpi_id):
    """
    POST /api/kpi-builder/<id>/evaluate/?bot_id=xxx
    Runs the formula evaluator and returns the computed numeric result.
    """
    from .models import CustomKPI
    from .services.kpi_evaluator_service import evaluate_formula

    bot_id = get_verified_bot_id(request, request.data.get('bot_id')) or request.query_params.get('bot_id')
    try:
        kpi = CustomKPI.objects.get(pk=kpi_id, bot_id=bot_id)
    except CustomKPI.DoesNotExist:
        return Response({'error': 'KPI not found.'}, status=status.HTTP_404_NOT_FOUND)

    result = evaluate_formula(bot_id, kpi.formula)
    return Response({
        'id': kpi.id,
        'name': kpi.name,
        'formula': kpi.formula,
        'unit': kpi.unit,
        'icon': kpi.icon,
        'color': kpi.color,
        **result,
    })


@api_view(['DELETE'])
def delete_custom_kpi(request, kpi_id):
    """
    DELETE /api/kpi-builder/<id>/?bot_id=xxx
    Soft-deletes (deactivates) a custom KPI.
    """
    from .models import CustomKPI
    bot_id = get_verified_bot_id(request, request.query_params.get('bot_id') or request.data.get('bot_id'))
    try:
        kpi = CustomKPI.objects.get(pk=kpi_id, bot_id=bot_id)
        kpi.is_active = False
        kpi.save()
        return Response({'message': f'KPI "{kpi.name}" deleted.'})
    except CustomKPI.DoesNotExist:
        return Response({'error': 'KPI not found.'}, status=status.HTTP_404_NOT_FOUND)

